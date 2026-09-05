# apps/administrador/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db import connection, IntegrityError
from django.db.models import Q
from django.utils import timezone
from django_fsm import can_proceed
from .models import (
    Usuario, Operario, Tarea,
    AsignacionTarea, Orden, Cliente, Incidencia, Inventario, Material, Producto, Factura,
    TIEMPOS_ESTANDAR_MINUTOS,
)
import json
import openpyxl
from datetime import datetime, date
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.conf import settings
import os
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.decorators import login_required_rol


# ── Decorador de protección por rol (centralizado en apps.core) ────
admin_required = login_required_rol(rol_esperado='administrador', session_key='usuario_id')


# ── Transiciones válidas de Orden, para editar el estado desde el admin ──
# (orden.estado, nuevo_estado) -> nombre del método @transition en el modelo.
# Mismo criterio que en apps/produccion/views.py: nunca se asigna
# orden.estado = texto_libre directamente.
TRANSICIONES_ORDEN = {
    ('Pendiente', 'Procesando'):  'marcar_en_produccion',
    ('Procesando', 'Enviado'):    'marcar_enviado',
    ('Enviado', 'Entregado'):     'marcar_entregado',
    ('Enviado', 'Procesando'):    'revertir_a_produccion',
    ('Pendiente', 'Cancelado'):   'cancelar',
    ('Procesando', 'Cancelado'):  'cancelar',
}


# ── Ubicaciones predefinidas del inventario ─────────────────
# Usadas tanto para validar/generar el <select> como para detectar,
# en la plantilla, cuándo una ubicación guardada es "personalizada"
# (es decir, no está en esta lista) y así preseleccionar "Otro".
UBICACIONES_PREDEFINIDAS = [
    'Bodega Principal',
    'Bodega Secundaria',
    'Bodega Formal',
    'Área de Producción',
    'Área de Despacho',
]


# ── Helpers para KPIs cross-módulo (antes vivían en produccion/views.py) ──
def _count(modelo, **filtros):
    try:
        return modelo.objects.filter(**filtros).count() if filtros else modelo.objects.count()
    except Exception:
        return 0


def _safe_import(ruta_modulo, nombre_modelo):
    try:
        modulo = __import__(ruta_modulo, fromlist=[nombre_modelo])
        return getattr(modulo, nombre_modelo)
    except Exception:
        return None


# ── Portal principal ─────────────────────────────────────────
@admin_required
def admin_portal(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

    total_usuarios = Usuario.objects.count()
    total_clientes = Cliente.objects.count()
    total_operarios = Operario.objects.filter(estado='activo').count()
    total_ordenes = Orden.objects.count()
    ordenes_pendientes = Orden.objects.filter(estado='Pendiente').count()
    tareas_pendientes = AsignacionTarea.objects.filter(estado='Pendiente').count()
    usuarios_pendientes = Usuario.objects.filter(estado='pendiente').count()

    ultimas_ordenes = Orden.objects.order_by('-fechaCreacion')[:5]
    ultimas_asignaciones = AsignacionTarea.objects.order_by('-fechaAsignacion')[:5]

    # ── KPIs exclusivos que venían de "Producción · Vista General" ──
    Produccion = _safe_import('apps.produccion.models', 'Produccion')
    Proveedor = _safe_import('apps.proveedores.models', 'Proveedor')

    ordenes_urgentes = Orden.objects.filter(prioridad='Urgente').count()
    incidencias_abiertas = (
        Incidencia.objects.filter(estado='Pendiente').count()
        + Incidencia.objects.filter(estado='En Progreso').count()
    )
    productos_catalogo = Producto.objects.count()
    produccion_activa = (
        _count(Produccion, estado='En Progreso') + _count(Produccion, estado='Pendiente')
        if Produccion else 0
    )
    total_proveedores = _count(Proveedor) if Proveedor else None

    # ── Actividad reciente unificada (Órdenes + Producción + Incidencias) ──
    actividad = []

    for o in Orden.objects.select_related('idCliente').order_by('-fechaCreacion')[:5]:
        cliente_nombre = '—'
        try:
            cliente_nombre = o.idCliente.empresa or o.idCliente.nombre or '—'
        except Exception:
            pass
        actividad.append({
            'icono': '🧾',
            'titulo': f'Orden #{o.idOrden} · {cliente_nombre}',
            'detalle': f'Estado: {o.estado} · Prioridad: {o.prioridad}',
            'fecha': str(o.fechaCreacion) if o.fechaCreacion else None,
            'estado': o.estado,
            'modulo': 'admin_ordenes',
        })

    if Produccion:
        try:
            for p in Produccion.objects.select_related('idProducto').order_by('-fechaInicio')[:5]:
                actividad.append({
                    'icono': '🧵',
                    'titulo': f'Producción: {p.idProducto.nombre}',
                    'detalle': f'{p.cantidadRequerida} unidades',
                    'fecha': str(p.fechaInicio) if p.fechaInicio else None,
                    'estado': p.estado,
                    'modulo': 'produccion_portal',
                })
        except Exception:
            pass

    try:
        for i in Incidencia.objects.order_by('-idIncidencia')[:5]:
            actividad.append({
                'icono': '⚠️',
                'titulo': getattr(i, 'tipoIncidencia', None) or (getattr(i, 'descripcion', '') or 'Incidencia')[:60],
                'detalle': f'Estado: {getattr(i, "estado", "—")}',
                'fecha': str(getattr(i, 'fechaGeneracion', '') or ''),
                'estado': getattr(i, 'estado', None),
                'modulo': 'admin_incidencias',
            })
    except Exception:
        pass

    actividad = [a for a in actividad if a['fecha']]
    actividad.sort(key=lambda a: a['fecha'], reverse=True)
    actividad_reciente = actividad[:10]

    # ── Alertas operativas ──
    alertas = []
    hoy = date.today()

    ordenes_retrasadas = Orden.objects.filter(
        fechaEntregaEstimada__lt=hoy
    ).exclude(estado__in=['Entregado', 'Cancelado']).count()
    if ordenes_retrasadas:
        alertas.append({
            'tipo': 'danger', 'icono': '⏰',
            'texto': f'{ordenes_retrasadas} orden(es) con entrega vencida',
            'modulo': 'admin_ordenes',
        })

    if ordenes_urgentes:
        alertas.append({
            'tipo': 'warning', 'icono': '🔥',
            'texto': f'{ordenes_urgentes} orden(es) marcadas como urgentes',
            'modulo': 'admin_ordenes',
        })

    if usuarios_pendientes:
        alertas.append({
            'tipo': 'info', 'icono': '👤',
            'texto': f'{usuarios_pendientes} usuario(s) esperando aprobación de rol',
            'modulo': 'admin_usuarios',
        })

    if incidencias_abiertas:
        alertas.append({
            'tipo': 'warning', 'icono': '⚠️',
            'texto': f'{incidencias_abiertas} incidencia(s) sin resolver',
            'modulo': 'admin_incidencias',
        })

    return render(request, 'administrador/admin_portal.html', {
        'usuario': usuario,
        'total_usuarios': total_usuarios,
        'total_clientes': total_clientes,
        'total_operarios': total_operarios,
        'total_ordenes': total_ordenes,
        'ordenes_pendientes': ordenes_pendientes,
        'tareas_pendientes': tareas_pendientes,
        'usuarios_pendientes': usuarios_pendientes,
        'ultimas_ordenes': ultimas_ordenes,
        'ultimas_asignaciones': ultimas_asignaciones,

        # ── Nuevos: exclusivos de producción trasladados al dashboard ──
        'ordenes_urgentes': ordenes_urgentes,
        'incidencias_abiertas': incidencias_abiertas,
        'productos_catalogo': productos_catalogo,
        'produccion_activa': produccion_activa,
        'total_proveedores': total_proveedores,
        'actividad_reciente': actividad_reciente,
        'alertas': json.dumps(alertas),
    })


# ── Usuarios ─────────────────────────────────────────────────
@admin_required
def usuarios_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    usuarios = Usuario.objects.all().order_by('-idUsuario')

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        usuarios = usuarios.filter(estado=estado_filtro)

    return render(request, 'administrador/usuarios_lista.html', {
        'usuario': usuario,
        'usuarios': usuarios,
        'estado_filtro': estado_filtro,
    })


@admin_required
def usuario_crear(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        correo = request.POST.get('correoElectronico')
        contrasena = request.POST.get('contrasena')
        telefono = request.POST.get('telefono', '')
        rol = request.POST.get('rol', 'cliente')

        if Usuario.objects.filter(correoElectronico=correo).exists():
            messages.error(request, 'Ya existe un usuario con ese correo.')
            return redirect('admin_usuario_crear')

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO usuarios
                    (nombre, apellido, correoElectronico, contrasena, telefono, rol, estado)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, [nombre, apellido, correo, make_password(contrasena),
                  telefono or None, rol, 'activo'])

            id_nuevo = cursor.lastrowid

            if rol == 'cliente':
                cursor.execute("""
                    INSERT INTO clientes (idUsuario, tipoCliente, nombre, correoElectronico, estado)
                    VALUES (%s, %s, %s, %s, %s)
                """, [id_nuevo, 'Natural', f'{nombre} {apellido}', correo, 'activo'])

            elif rol == 'operario':
                especialidad = request.POST.get('especialidad', 'General')
                cursor.execute("""
                    INSERT INTO operarios (idUsuario, especialidad, fechaIngreso, estado)
                    VALUES (%s, %s, CURDATE(), %s)
                """, [id_nuevo, especialidad, 'activo'])

        messages.success(request, f'Usuario {nombre} {apellido} creado correctamente.')
        return redirect('admin_usuarios')

    return render(request, 'administrador/usuario_form.html', {
        'usuario': usuario,
        'accion': 'Crear',
    })


@admin_required
def usuario_editar(request, idUsuario):
    usuario_admin = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    usuario_editar_obj = Usuario.objects.get(idUsuario=idUsuario)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        correo = request.POST.get('correoElectronico')
        telefono = request.POST.get('telefono', '')
        rol = request.POST.get('rol')
        estado = request.POST.get('estado')

        usuario_editar_obj.nombre = nombre
        usuario_editar_obj.apellido = apellido
        usuario_editar_obj.correoElectronico = correo
        usuario_editar_obj.telefono = telefono or None
        usuario_editar_obj.rol = rol
        usuario_editar_obj.estado = estado
        usuario_editar_obj.save()

        with connection.cursor() as cursor:
            if rol == 'cliente':
                cursor.execute(
                    "SELECT idCliente FROM clientes WHERE idUsuario = %s", [idUsuario]
                )
                if cursor.fetchone() is None:
                    cursor.execute("""
                        INSERT INTO clientes (idUsuario, tipoCliente, nombre, correoElectronico, estado)
                        VALUES (%s, %s, %s, %s, %s)
                    """, [idUsuario, 'Natural', f'{nombre} {apellido}', correo, 'activo'])

            elif rol == 'operario':
                cursor.execute(
                    "SELECT idOperario FROM operarios WHERE idUsuario = %s", [idUsuario]
                )
                if cursor.fetchone() is None:
                    especialidad = request.POST.get('especialidad', 'General')
                    cursor.execute("""
                        INSERT INTO operarios (idUsuario, especialidad, fechaIngreso, estado)
                        VALUES (%s, %s, CURDATE(), %s)
                    """, [idUsuario, especialidad, 'activo'])

        messages.success(request, f'Usuario {nombre} actualizado correctamente.')
        return redirect('admin_usuarios')

    return render(request, 'administrador/usuario_form.html', {
        'usuario': usuario_admin,
        'usuario_editar': usuario_editar_obj,
        'accion': 'Editar',
    })


@admin_required
def usuario_eliminar(request, idUsuario):
    if request.method == 'POST':
        usuario_obj = get_object_or_404(Usuario, idUsuario=idUsuario)

        # Validación de seguridad: Solo permitir eliminar si es operario
        if usuario_obj.rol != 'operario':
            messages.error(request, '⚠️ No está permitido eliminar usuarios con rol diferente a Operario.')
            return redirect('admin_usuarios')

        nombre = f'{usuario_obj.nombre} {usuario_obj.apellido}'
        usuario_obj.delete()
        messages.success(request, f'✅ Usuario operario {nombre} eliminado correctamente.')

    return redirect('admin_usuarios')


# ── Órdenes ──────────────────────────────────────────────────
@admin_required
def ordenes_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    ordenes = Orden.objects.all().order_by('-fechaCreacion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    # Soporta también el filtro por prioridad que ahora enlazan las
    # tarjetas KPI de "Órdenes Urgentes" del dashboard.
    prioridad_filtro = request.GET.get('prioridad', '')
    if prioridad_filtro:
        ordenes = ordenes.filter(prioridad=prioridad_filtro)

    return render(request, 'administrador/ordenes_lista.html', {
        'usuario': usuario,
        'ordenes': ordenes,
        'estado_filtro': estado_filtro,
        'buscar_filtro': buscar_filtro,
        'prioridad_filtro': prioridad_filtro,
        # Para que la plantilla sepa qué opciones de estado mostrar en el
        # <select> de edición según el estado actual de cada orden.
        'transiciones_orden': TRANSICIONES_ORDEN,
    })


@admin_required
def orden_editar(request, idOrden):
    if request.method == 'POST':
        orden = get_object_or_404(Orden, pk=idOrden)
        cantidad = request.POST.get('cantidad')
        precio_unitario = request.POST.get('precio_unitario')
        fecha_entrega = request.POST.get('fecha_entrega')
        prioridad = request.POST.get('prioridad')
        nuevo_estado = request.POST.get('estado')

        orden.cantidad = int(cantidad) if cantidad and cantidad.strip() else None
        orden.precioUnitario = float(precio_unitario) if precio_unitario and precio_unitario.strip() else None
        orden.fechaEntregaEstimada = fecha_entrega if fecha_entrega and fecha_entrega.strip() else None
        orden.prioridad = prioridad

        if nuevo_estado and nuevo_estado != orden.estado:
            metodo_nombre = TRANSICIONES_ORDEN.get((orden.estado, nuevo_estado))
            if not metodo_nombre:
                messages.error(
                    request,
                    f'No se puede pasar la orden #{idOrden} de "{orden.estado}" a "{nuevo_estado}".'
                )
                return redirect('admin_ordenes')
            metodo = getattr(orden, metodo_nombre)
            if not can_proceed(metodo):
                messages.error(request, f'Transición "{metodo_nombre}" no permitida en este momento.')
                return redirect('admin_ordenes')
            metodo()

        orden.save()
        messages.success(request, f'La orden #{idOrden} se ha modificado con éxito.')

    return redirect('admin_ordenes')


@admin_required
def orden_eliminar(request, idOrden):
    try:
        orden = get_object_or_404(Orden, pk=idOrden)
        orden.delete()
        messages.success(request, f'La orden #{idOrden} se eliminó correctamente.')
    except Exception as e:
        messages.error(request, f'Error al intentar eliminar la orden: {str(e)}')
    return redirect('admin_ordenes')


# ── Helpers de fechas ────────────────────────────────────────
def _parsear_fecha(valor):
    """Convierte 'YYYY-MM-DD' (input type=date) a date, o None si viene vacío."""
    if not valor or not valor.strip():
        return None
    return datetime.strptime(valor.strip(), '%Y-%m-%d').date()


# ── Tareas ───────────────────────────────────────────────────
@admin_required
def tarea_asignar(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    operarios = Operario.objects.filter(estado='activo').select_related('idUsuario')
    tareas = Tarea.objects.all()
    # Solo mostramos órdenes que aún tiene sentido producir
    ordenes = Orden.objects.exclude(estado__in=['Cancelado', 'Entregado']) \
        .select_related('idCliente') \
        .order_by('-fechaCreacion')

    if request.method == 'POST':
        id_tarea = request.POST.get('tarea')
        tarea_personalizada = request.POST.get('tarea_personalizada', '').strip()
        proceso_personalizado = request.POST.get('proceso_personalizado', '').strip()
        ids_operarios = request.POST.getlist('operarios')
        id_orden = request.POST.get('orden')
        descripcion = request.POST.get('descripcion')
        fecha_inicio = request.POST.get('fechaInicio')
        fecha_limite = request.POST.get('fechaLimite')
        prioridad = request.POST.get('prioridad', 'Media')
        tipo_prenda = request.POST.get('tipoPrenda')
        cantidad = request.POST.get('cantidadPrendas')
        horas_estimadas = request.POST.get('horasEstimadas')

        try:
            if not ids_operarios:
                messages.error(request, 'Debes seleccionar al menos un operario.')
                return redirect('admin_tarea_asignar')

            if id_tarea == 'otra':
                if not tarea_personalizada:
                    messages.error(request, 'Por favor, ingresa el nombre de la tarea personalizada.')
                    return redirect('admin_tarea_asignar')
                if not proceso_personalizado:
                    messages.error(request, 'Por favor, ingresa el proceso/categoría de la tarea.')
                    return redirect('admin_tarea_asignar')

                tarea = Tarea.objects.create(
                    nombreTarea=tarea_personalizada,
                    descripcionTarea=descripcion or f'Tarea personalizada: {tarea_personalizada}',
                    proceso=proceso_personalizado,
                    complejidad='media'
                )
                mensaje_tarea = f'✓ Tarea personalizada "{tarea_personalizada}" creada. '
            else:
                try:
                    tarea = Tarea.objects.get(idTarea=id_tarea)
                    mensaje_tarea = ''
                except Tarea.DoesNotExist:
                    messages.error(request, 'La tarea seleccionada no existe.')
                    return redirect('admin_tarea_asignar')

            operarios_seleccionados = list(
                Operario.objects.select_related('idUsuario').filter(idOperario__in=ids_operarios)
            )
            if len(operarios_seleccionados) != len(ids_operarios):
                messages.error(request, 'Uno o más operarios seleccionados no existen.')
                return redirect('admin_tarea_asignar')

            ESTADOS_ACTIVOS = ['Pendiente', 'En Progreso']
            ocupados = []
            for operario in operarios_seleccionados:
                tiene_activa = AsignacionTarea.objects.filter(
                    idOperario=operario,
                    estado__in=ESTADOS_ACTIVOS
                ).exists()
                if tiene_activa:
                    ocupados.append(f'{operario.idUsuario.nombre} {operario.idUsuario.apellido}')

            if ocupados:
                messages.error(
                    request,
                    'No se puede asignar: los siguientes operarios ya tienen una tarea activa '
                    '(Pendiente o En Progreso) → ' + ', '.join(ocupados) +
                    '. Un operario solo puede tener una tarea activa a la vez.'
                )
                return redirect('admin_tarea_asignar')

            orden = None
            if id_orden:
                try:
                    orden = Orden.objects.get(idOrden=id_orden)
                except Orden.DoesNotExist:
                    messages.error(request, 'La orden seleccionada no existe.')
                    return redirect('admin_tarea_asignar')

            cantidad_int = int(cantidad) if cantidad and cantidad.strip() else None

            fecha_inicio_dt = _parsear_fecha(fecha_inicio)
            fecha_limite_dt = _parsear_fecha(fecha_limite)

            if fecha_inicio_dt and fecha_inicio_dt < date.today():
                messages.error(request, 'La fecha de inicio no puede ser anterior a hoy.')
                return redirect('admin_tarea_asignar')

            if fecha_limite_dt and fecha_limite_dt < date.today():
                messages.error(request, 'La fecha límite no puede ser anterior a hoy.')
                return redirect('admin_tarea_asignar')

            if fecha_inicio_dt and fecha_limite_dt and fecha_limite_dt < fecha_inicio_dt:
                messages.error(
                    request,
                    'La fecha límite no puede ser anterior a la fecha de inicio.'
                )
                return redirect('admin_tarea_asignar')

            if (not horas_estimadas or not horas_estimadas.strip()) and tipo_prenda and cantidad_int:
                minutos_unidad = TIEMPOS_ESTANDAR_MINUTOS.get(tipo_prenda, 0)
                horas_calculadas = round((cantidad_int * minutos_unidad) / 60, 2)
            else:
                horas_calculadas = float(horas_estimadas) if horas_estimadas and horas_estimadas.strip() else 0.5

            asignaciones_creadas = []
            for operario in operarios_seleccionados:
                asignacion = AsignacionTarea.objects.create(
                    idTarea=tarea,
                    idOperario=operario,
                    idOrden=orden,
                    descripcion=descripcion,
                    fechaInicio=fecha_inicio_dt,
                    fechaLimite=fecha_limite_dt,
                    prioridad=prioridad,
                    tipoPrenda=tipo_prenda or None,
                    cantidadPrendas=cantidad_int,
                    horasEstimadas=horas_calculadas,
                    estado='Pendiente'
                )
                asignaciones_creadas.append(asignacion)

            nombres = ', '.join(
                f'{op.idUsuario.nombre} {op.idUsuario.apellido}' for op in operarios_seleccionados
            )
            messages.success(
                request,
                f'{mensaje_tarea}Se crearon {len(asignaciones_creadas)} asignación(es) correctamente. '
                f'Tarea asignada a: {nombres}.'
            )
            return redirect('admin_tareas')

        except Exception as e:
            messages.error(request, f'Error al asignar tarea: {str(e)}')
            return redirect('admin_tarea_asignar')

    return render(request, 'administrador/tarea_asignar.html', {
        'usuario': usuario,
        'operarios': operarios,
        'tareas': tareas,
        'ordenes': ordenes,
        'tiempos_estandar': TIEMPOS_ESTANDAR_MINUTOS,
    })


@admin_required
def tareas_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    asignaciones = AsignacionTarea.objects.select_related(
        'idTarea', 'idOperario__idUsuario', 'idOrden'
    ).order_by('-fechaAsignacion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        asignaciones = asignaciones.filter(
            Q(idTarea__nombreTarea__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__nombre__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__apellido__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        asignaciones = asignaciones.filter(estado=estado_filtro)

    ordenes = Orden.objects.exclude(estado__in=['Cancelado', 'Entregado']) \
        .select_related('idCliente') \
        .order_by('-fechaCreacion')

    return render(request, 'administrador/tareas_lista.html', {
        'usuario': usuario,
        'asignaciones': asignaciones,
        'buscar_filtro': buscar_filtro,
        'estado_filtro': estado_filtro,
        'ordenes': ordenes,
    })


@admin_required
def tarea_editar(request, idAsignacion):
    asignacion = get_object_or_404(AsignacionTarea, pk=idAsignacion)

    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_limite = request.POST.get('fecha_limite')
        estado = request.POST.get('estado')
        prioridad = request.POST.get('prioridad')
        tipo_prenda = request.POST.get('tipoPrenda')
        cantidad_prendas = request.POST.get('cantidadPrendas')
        horas_estimadas = request.POST.get('horas_estimadas')
        id_orden = request.POST.get('orden')

        try:
            if descripcion is not None:
                asignacion.descripcion = descripcion

            fecha_inicio_dt = _parsear_fecha(fecha_inicio)
            fecha_limite_dt = _parsear_fecha(fecha_limite)

            if fecha_inicio_dt:
                asignacion.fechaInicio = fecha_inicio_dt
            asignacion.fechaLimite = fecha_limite_dt

            if estado:
                asignacion.estado = estado
            if prioridad:
                asignacion.prioridad = prioridad

            asignacion.tipoPrenda = tipo_prenda or None
            asignacion.cantidadPrendas = int(cantidad_prendas) if cantidad_prendas and cantidad_prendas.strip() else None

            if horas_estimadas and horas_estimadas.strip():
                asignacion.horasEstimadas = float(horas_estimadas)

            if id_orden:
                try:
                    asignacion.idOrden = Orden.objects.get(idOrden=id_orden)
                except Orden.DoesNotExist:
                    messages.error(request, 'La orden seleccionada no existe.')
                    return redirect('admin_tareas')
            else:
                asignacion.idOrden = None

            asignacion.save()

            # ── Igual que en apps/operarios/views.py: si esta tarea
            # pertenece a un lote de producción, recalculamos su avance.
            id_produccion = asignacion.idTarea.idProduccion
            if id_produccion:
                from apps.produccion.services import recalcular_produccion_desde_tareas
                recalcular_produccion_desde_tareas(id_produccion)

            messages.success(request, f'Asignación #{idAsignacion} actualizada correctamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar la asignación: {str(e)}')

    return redirect('admin_tareas')


@admin_required
def tarea_eliminar(request, idAsignacion):
    if request.method == 'POST':
        try:
            asignacion = get_object_or_404(AsignacionTarea, pk=idAsignacion)
            asignacion.delete()
            messages.success(request, f'Asignación #{idAsignacion} eliminada correctamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la asignación: {str(e)}')
    return redirect('admin_tareas')


# ── Incidencias ──────────────────────────────────────────────
@admin_required
def incidencias_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    incidencias = Incidencia.objects.all().order_by('-fechaGeneracion')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        incidencias = incidencias.filter(
            Q(tipoIncidencia__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__nombre__icontains=buscar_filtro) |
            Q(idOperario__idUsuario__apellido__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        incidencias = incidencias.filter(estado=estado_filtro)

    MESES_ES = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
    }
    hoy = date.today()
    periodo_actual = f'{MESES_ES[hoy.month]} {hoy.year}'

    return render(request, 'administrador/incidencias_lista.html', {
        'usuario': usuario,
        'incidencias': incidencias,
        'buscar_filtro': buscar_filtro,
        'estado_filtro': estado_filtro,
        'periodo_actual': periodo_actual,
    })


@admin_required
def incidencia_editar(request, idIncidencia):
    if request.method == 'POST':
        incidencia = get_object_or_404(Incidencia, pk=idIncidencia)
        incidencia.tipoIncidencia = request.POST.get('tipoIncidencia')
        incidencia.descripcion = request.POST.get('descripcion')
        periodo_evaluado = (request.POST.get('periodoEvaluado') or '').strip()
        if not periodo_evaluado:
            MESES_ES = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
            }
            hoy = date.today()
            periodo_evaluado = f'{MESES_ES[hoy.month]} {hoy.year}'
        incidencia.periodoEvaluado = periodo_evaluado
        incidencia.estado = request.POST.get('estado')
        fecha_revision = request.POST.get('fechaRevision')
        incidencia.fechaRevision = fecha_revision if fecha_revision and fecha_revision.strip() else None

        nueva_respuesta = (request.POST.get('respuesta') or '').strip() or None
        if nueva_respuesta != incidencia.respuesta:
            incidencia.respuestaLeida = False
        incidencia.respuesta = nueva_respuesta

        incidencia.save()
        messages.success(request, f'Incidencia #{idIncidencia} actualizada correctamente.')
    return redirect('admin_incidencias')


@admin_required
def incidencia_eliminar(request, idIncidencia):
    if request.method == 'POST':
        try:
            incidencia = get_object_or_404(Incidencia, pk=idIncidencia)
            incidencia.delete()
            messages.success(request, f'Incidencia #{idIncidencia} eliminada correctamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la incidencia: {str(e)}')
    return redirect('admin_incidencias')


# ── Facturas ─────────────────────────────────────────────────
@admin_required
def facturas_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])
    facturas = Factura.objects.select_related('idCliente', 'idOrden').order_by('-fechaEmision')

    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        facturas = facturas.filter(
            Q(numeroFactura__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro) |
            Q(idOrden__idOrden__icontains=buscar_filtro)
        )

    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        facturas = facturas.filter(estado=estado_filtro)

    return render(request, 'administrador/facturas_lista.html', {
        'usuario': usuario,
        'facturas': facturas,
        'buscar_filtro': buscar_filtro,
        'estado_filtro': estado_filtro,
        'seccion_activa': 'facturas',
    })


@admin_required
def factura_marcar_pagada(request, idFactura):
    if request.method == 'POST':
        try:
            factura = get_object_or_404(Factura, pk=idFactura)
            if factura.estado == 'Pagada':
                messages.warning(request, f'La factura {factura.numeroFactura} ya estaba marcada como pagada.')
            else:
                factura.estado = 'Pagada'
                factura.fechaPago = timezone.now()
                factura.save()
                messages.success(request, f'Factura {factura.numeroFactura} marcada como pagada.')
        except Exception as e:
            messages.error(request, f'Error al actualizar la factura: {str(e)}')
    return redirect('admin_facturas')


@admin_required
def factura_descargar(request, idFactura):
    """Descarga de PDF para admin — sin el filtro de 'factura propia del cliente'."""
    factura = get_object_or_404(Factura, pk=idFactura)
    ruta = os.path.join(settings.MEDIA_ROOT, factura.rutaPDF)
    if not os.path.exists(ruta):
        raise Http404('El archivo de la factura no fue encontrado.')
    return FileResponse(
        open(ruta, 'rb'),
        as_attachment=True,
        filename=os.path.basename(ruta)
    )


# ── Módulos / Placeholders externos ──────────────────────────
@admin_required
def produccion_placeholder(request):
    return redirect('produccion_portal')


# ── Exportar Órdenes a Excel ──────────────────────────────────
@admin_required
def exportar_ordenes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes HebraTech"

    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_REGULAR = Font(name="Calibri", size=11)
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    headers = ["ID Orden", "Cliente / Empresa", "Fecha Creación", "Entrega Estimada",
               "Cantidad", "Precio Unitario", "Total", "Prioridad", "Estado"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ordenes = Orden.objects.all().select_related('idCliente__idUsuario').order_by('-fechaCreacion')
    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )
    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    for idx, orden in enumerate(ordenes, start=2):
        cliente_nombre = orden.idCliente.empresa or orden.idCliente.nombre or f"Cliente #{orden.idCliente.idCliente}"
        ws.append([
            orden.idOrden, cliente_nombre,
            orden.fechaCreacion.strftime('%Y-%m-%d') if orden.fechaCreacion else "",
            orden.fechaEntregaEstimada.strftime('%Y-%m-%d') if orden.fechaEntregaEstimada else "",
            orden.cantidad or 0,
            float(orden.precioUnitario) if orden.precioUnitario else 0,
            f"=E{idx}*F{idx}",
            orden.prioridad, orden.estado
        ])
        is_zebra = (idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = FONT_REGULAR
            cell.border = THIN_BORDER
            if is_zebra:
                cell.fill = ZEBRA_FILL
            if col_idx in [1, 3, 4, 8, 9]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="HebraTech_Reporte_Ordenes.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_ordenes_pdf(request):
    ordenes = Orden.objects.all().select_related('idCliente__idUsuario').order_by('-fechaCreacion')
    buscar_filtro = request.GET.get('buscar', '')
    if buscar_filtro:
        ordenes = ordenes.filter(
            Q(idOrden__icontains=buscar_filtro) |
            Q(idCliente__nombre__icontains=buscar_filtro) |
            Q(idCliente__empresa__icontains=buscar_filtro)
        )
    estado_filtro = request.GET.get('estado', '')
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    html_string = render_to_string('administrador/ordenes_pdf.html', {'ordenes': ordenes})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="HebraTech_Reporte_Ordenes.pdf"'
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=500)
    return response


# ── Inventario y Materiales ──────────────────────────────────
@admin_required
def inventario_lista(request):
    usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

    buscar = request.GET.get('buscar', '').strip()

    inventario_list = Inventario.objects.all().select_related('producto')
    materiales_list = Material.objects.all()
    productos_list = Producto.objects.all()

    if buscar:
        inventario_list = inventario_list.filter(
            Q(producto__nombre__icontains=buscar) |
            Q(idInventario__icontains=buscar) |
            Q(ubicacion__icontains=buscar)
        )

        materiales_list = materiales_list.filter(
            Q(nombreMaterial__icontains=buscar) |
            Q(descripcion__icontains=buscar) |
            Q(idMaterial__icontains=buscar)
        )

    context = {
        'usuario': usuario,
        'inventario_list': inventario_list,
        'total_items': inventario_list.count(),
        'materiales_list': materiales_list,
        'total_materiales': materiales_list.count(),
        'productos_list': productos_list,
        'buscar_filtro': buscar,
        'ubicaciones_predefinidas': UBICACIONES_PREDEFINIDAS,
    }
    return render(request, 'administrador/inventario_lista.html', context)


# ── CRUD: MATERIALES ─────────────────────────────────────────

@admin_required
def crear_material(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombreMaterial')
        descripcion = request.POST.get('descripcion')
        stock_actual = request.POST.get('stockActual') or 0
        stock_minimo = request.POST.get('stockMinimo') or 0
        unidad = request.POST.get('unidadBase')
        costo = request.POST.get('costoUnitario') or 0

        Material.objects.create(
            nombreMaterial=nombre,
            descripcion=descripcion,
            stockActual=stock_actual,
            stockMinimo=stock_minimo,
            unidadBase=unidad,
            costoUnitario=costo
        )
        messages.success(request, f"Material '{nombre}' creado con éxito.")
    return redirect('admin_inventario')


@admin_required
def editar_material(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        material.nombreMaterial = request.POST.get('nombreMaterial')
        material.descripcion = request.POST.get('descripcion')
        material.stockActual = request.POST.get('stockActual')
        material.stockMinimo = request.POST.get('stockMinimo')
        material.unidadBase = request.POST.get('unidadBase')
        material.costoUnitario = request.POST.get('costoUnitario')

        material.save()
        messages.success(request, f"Material '{material.nombreMaterial}' actualizado correctamente.")
    return redirect('admin_inventario')


@admin_required
def eliminar_material(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        nombre = material.nombreMaterial
        material.delete()
        messages.success(request, f"Material '{nombre}' eliminado correctamente.")
    return redirect('admin_inventario')


# ── Helper: resolver ubicación (predefinida u "Otro") ────────
def _resolver_ubicacion(request):
    """
    Si el <select name="ubicacion"> viene con value="otro", usa el texto libre
    escrito en el input "ubicacion_personalizada". En cualquier otro caso,
    devuelve tal cual el valor seleccionado (o None si viene vacío).
    """
    ubicacion = request.POST.get('ubicacion')
    if ubicacion == 'otro':
        ubicacion = (request.POST.get('ubicacion_personalizada') or '').strip() or None
    return ubicacion or None


# ── CRUD: INVENTARIO (PRODUCTOS) ─────────────────────────────
@admin_required
def crear_inventario(request):
    if request.method == 'POST':
        nombre_producto = request.POST.get('nombre_producto', '').strip()

        if not nombre_producto:
            messages.error(request, "Debes ingresar el nombre del producto.")
            return redirect('admin_inventario')

        producto, creado = Producto.objects.get_or_create(
            nombre=nombre_producto,
            defaults={
                'descripcion': f'Producto registrado desde inventario ({nombre_producto})',
                'precio': 0,
                'categoria': 'Sin categoría',
                'estado': 'activo',
            }
        )

        if not creado and Inventario.objects.filter(producto=producto).exists():
            messages.error(
                request,
                f"Ya existe un registro de inventario para el producto '{producto.nombre}'. Edítalo en lugar de crear uno nuevo."
            )
            return redirect('admin_inventario')

        try:
            cant_disponible = int(request.POST.get('cantidadDisponible') or 0)
            min_definido = int(request.POST.get('minimoDefinido') or 0)
        except ValueError:
            messages.error(request, "La cantidad disponible y el mínimo definido deben ser números enteros.")
            return redirect('admin_inventario')

        if cant_disponible < 0 or min_definido < 0:
            messages.error(request, "Los valores de stock no pueden ser negativos.")
            return redirect('admin_inventario')

        if cant_disponible < min_definido:
            messages.error(request, "El stock actual no puede ser menor al mínimo definido.")
            return redirect('admin_inventario')

        try:
            cant_ingresada = int(request.POST.get('cantidadIngresada') or 0)
        except ValueError:
            messages.error(request, "La cantidad ingresada debe ser un número entero.")
            return redirect('admin_inventario')

        if cant_ingresada < 0:
            messages.error(request, "La cantidad ingresada no puede ser negativa.")
            return redirect('admin_inventario')

        nivel_stock_input = request.POST.get('nivelStock')
        try:
            nivel_stock = int(nivel_stock_input)
        except (ValueError, TypeError):
            nivel_stock = cant_disponible - min_definido

        unidades = request.POST.get('unidades') or 'Unidades'
        ubicacion = _resolver_ubicacion(request)
        cant_egresada = 0
        fecha_ingreso = request.POST.get('fechaIngreso') or timezone.now().date()
        fecha_salida = request.POST.get('fechaSalida') or None

        try:
            Inventario.objects.create(
                producto=producto,
                cantidadDisponible=cant_disponible,
                minimoDefinido=min_definido,
                nivelStock=nivel_stock,
                unidades=unidades,
                ubicacion=ubicacion,
                cantidadIngresada=cant_ingresada,
                cantidadEgresada=cant_egresada,
                fechaIngreso=fecha_ingreso,
                fechaSalida=fecha_salida if fecha_salida else None
            )
            messages.success(request, f"Registro de inventario para '{producto.nombre}' creado exitosamente.")
        except IntegrityError:
            messages.error(
                request,
                f"Ya existe un registro de inventario para el producto '{producto.nombre}'."
            )

    return redirect('admin_inventario')


@admin_required
def editar_inventario(request, pk):
    item = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        producto_id = request.POST.get('producto')
        nuevo_producto = get_object_or_404(Producto, pk=producto_id)

        if nuevo_producto.pk != item.producto.pk and \
                Inventario.objects.filter(producto=nuevo_producto).exclude(pk=item.pk).exists():
            messages.error(
                request,
                f"El producto '{nuevo_producto.nombre}' ya tiene un registro de inventario."
            )
            return redirect('admin_inventario')

        try:
            cant_disponible = int(request.POST.get('cantidadDisponible') or 0)
            min_definido = int(request.POST.get('minimoDefinido') or 0)
        except ValueError:
            messages.error(request, "La cantidad disponible y el mínimo definido deben ser números enteros.")
            return redirect('admin_inventario')

        if cant_disponible < 0 or min_definido < 0:
            messages.error(request, "Los valores de stock no pueden ser negativos.")
            return redirect('admin_inventario')

        if cant_disponible < min_definido:
            messages.error(request, "El stock actual no puede ser menor al mínimo definido.")
            return redirect('admin_inventario')

        try:
            cant_ingresada = int(request.POST.get('cantidadIngresada') or 0)
            cant_egresada = int(request.POST.get('cantidadEgresada') or 0)
        except ValueError:
            messages.error(request, "Las cantidades ingresada/egresada deben ser números enteros.")
            return redirect('admin_inventario')

        if cant_ingresada < 0 or cant_egresada < 0:
            messages.error(request, "Las cantidades ingresada/egresada no pueden ser negativas.")
            return redirect('admin_inventario')

        item.producto = nuevo_producto
        item.cantidadDisponible = cant_disponible
        item.minimoDefinido = min_definido

        nivel_stock_input = request.POST.get('nivelStock')
        try:
            item.nivelStock = int(nivel_stock_input)
        except (ValueError, TypeError):
            item.nivelStock = cant_disponible - min_definido

        item.unidades = request.POST.get('unidades')
        item.ubicacion = _resolver_ubicacion(request)
        item.cantidadIngresada = cant_ingresada
        item.cantidadEgresada = cant_egresada

        fecha_salida = request.POST.get('fechaSalida')
        item.fechaSalida = fecha_salida if fecha_salida else None

        try:
            item.save()
            messages.success(request, f"Registro de inventario #{item.idInventario} actualizado.")
        except IntegrityError:
            messages.error(
                request,
                f"El producto '{nuevo_producto.nombre}' ya tiene un registro de inventario."
            )

    return redirect('admin_inventario')


@admin_required
def eliminar_inventario(request, pk):
    item = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        id_inv = item.idInventario
        item.delete()
        messages.success(request, f"Registro de inventario #{id_inv} eliminado.")
    return redirect('admin_inventario')


# ── Perfil de Usuario ────────────────────────────────────────

@admin_required
@require_POST
def editar_perfil(request):
    if request.method == 'POST':
        try:
            usuario = Usuario.objects.get(idUsuario=request.session['usuario_id'])

            nombre = request.POST.get('nombre', '').strip()
            apellido = request.POST.get('apellido', '').strip()
            correo = request.POST.get('email', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            pass1 = request.POST.get('password1', '').strip()
            foto = request.FILES.get('foto')

            if not correo or not nombre:
                return JsonResponse({'success': False, 'message': 'Nombre y Correo electrónico son obligatorios.'}, status=400)

            usuario.nombre = nombre
            usuario.apellido = apellido
            usuario.correoElectronico = correo
            usuario.telefono = telefono or None

            if foto:
                usuario.fotoPerfil = foto

            if pass1:
                usuario.contrasena = make_password(pass1)

            usuario.save()

            request.session['usuario_nombre'] = usuario.nombre

            return JsonResponse({'success': True, 'message': 'Perfil actualizado correctamente.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Error al actualizar: {str(e)}"}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
